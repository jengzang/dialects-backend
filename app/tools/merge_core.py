"""
Merge工具核心逻辑
从 app/service/utils/merge/wordsheet_merge.py 提取
保持原有逻辑完全不变
"""
import os
import openpyxl
from openpyxl import Workbook
from openpyxl.comments import Comment
from collections import defaultdict
from common.constants import col_map


def load_reference_file(reference_path):
    """
    加载参考表
    原函数：wordsheet_merge.py 第9-55行

    Args:
        reference_path: 参考表文件路径

    Returns:
        参考字符列表
    """
    ref_wb = openpyxl.load_workbook(reference_path, read_only=True, keep_vba=True)

    def get_single_characters(sheet):
        headers = [cell.value for cell in next(sheet.iter_rows(min_row=1, max_row=1))]
        char_col_indices = [i for i, h in enumerate(headers) if h in ("單字", "单字")]

        if len(char_col_indices) == 0:
            raise ValueError("未找到單字或单字列")
        if len(char_col_indices) > 1:
            raise ValueError("表中同时存在單字与单字列，请删除其中一个以避免歧义。")

        char_col_idx = char_col_indices[0]
        chars = []
        for row in sheet.iter_rows(min_row=2):
            value = row[char_col_idx].value
            if value:
                chars.append(str(value))
        return chars

    sheetnames = ref_wb.sheetnames
    main_sheet = None
    supplement_sheet = None

    if "主表" in sheetnames:
        main_sheet = ref_wb["主表"]
    else:
        main_sheet = ref_wb[sheetnames[0]]

    # 定義可能的名稱，按優先級排序（先繁後簡）
    targets = ["補充表", "补充表"]
    supplement_sheet = None

    for name in targets:
        if name in ref_wb.sheetnames:
            supplement_sheet = ref_wb[name]
            break  # 找到第一個就跳出迴圈
    if not supplement_sheet:
        print("警告：找不到補充表（繁體或簡體）")

    main_chars = get_single_characters(main_sheet)
    main_set = set(main_chars)

    if supplement_sheet:
        supplement_chars = get_single_characters(supplement_sheet)
        additional_chars = [char for char in supplement_chars if char not in main_set]
    else:
        additional_chars = []

    final_chars = main_chars.copy()
    if additional_chars:
        final_chars.append("-")
        final_chars.extend(additional_chars)

    return final_chars


def merge_excel_files(reference_chars, files):
    """
    合并Excel文件（优化版：减少重复计算，使用 constants.col_map）
    原函数：wordsheet_merge.py 第70-132行

    Args:
        reference_chars: 参考字符列表
        files: 待合并文件路径列表

    Returns:
        (merged_data, comments_data): 合并后的数据和批注数据
    """
    merged_data = {char: [""] * len(files) for char in reference_chars}  # 存放每个表的 syllable
    comments_data = {char: [[] for _ in range(len(files))] for char in reference_chars}  # 存放每个表的批注

    # 使用 constants.col_map 并转换为集合以加速查找
    column_aliases = {
        'phrase': set(col_map['漢字']),
        'syllable': set(col_map['音標']),
        'notes': set(col_map['解釋'])
    }

    def find_column_index(header, targets):
        """在 header 中找出第一个匹配的 targets 项，并返回其索引（使用集合查找）"""
        for i, name in enumerate(header):
            if name in targets:
                return i
        return None

    for file_index, file in enumerate(files):
        wb = openpyxl.load_workbook(file, read_only=True)
        for ws in wb.worksheets:
            header = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]

            phrase_col = find_column_index(header, column_aliases['phrase'])
            syllable_col = find_column_index(header, column_aliases['syllable'])
            notes_col = find_column_index(header, column_aliases['notes'])

            # 跳过没有核心列的表
            if phrase_col is None or syllable_col is None:
                continue

            # 统计 phrase 重复数
            phrase_count = defaultdict(int)
            for row in ws.iter_rows(min_row=2):
                phrase = row[phrase_col].value
                phrase_count[phrase] += 1

            # 遍历数据行
            for row in ws.iter_rows(min_row=2):
                phrase = row[phrase_col].value
                syllable = row[syllable_col].value
                note = row[notes_col].value if notes_col is not None else None

                if phrase in merged_data:
                    if merged_data[phrase][file_index]:
                        merged_data[phrase][file_index] += f";{syllable}"
                        if note and phrase_count[phrase] > 1:
                            comments_data[phrase][file_index].append(note)
                    else:
                        merged_data[phrase][file_index] = syllable
                        if note and phrase_count[phrase] > 1:
                            comments_data[phrase][file_index].append(note)

    # 清理重复内容（优化：只在有分号时才拆分）
    for phrase in merged_data:
        for i in range(len(merged_data[phrase])):
            entry = merged_data[phrase][i]
            if entry and ";" in entry:
                parts = [part.strip() for part in entry.split(";")]
                if all(p == parts[0] for p in parts):
                    merged_data[phrase][i] = parts[0]

    return merged_data, comments_data


def get_file_name(file_path):
    """
    获取文件名（不带路径和扩展名）
    原函数：wordsheet_merge.py 第136-137行

    Args:
        file_path: 文件路径

    Returns:
        文件名（不带扩展名）
    """
    return os.path.splitext(os.path.basename(file_path))[0]


def create_new_workbook(reference_chars, merged_data, comments_data, file_names):
    """
    创建新的Excel工作簿用于存储合并结果
    原函数：wordsheet_merge.py 第141-170行

    Args:
        reference_chars: 参考字符列表
        merged_data: 合并后的数据
        comments_data: 批注数据
        file_names: 文件名列表

    Returns:
        openpyxl.Workbook对象
    """
    new_wb = Workbook()
    new_ws = new_wb.active
    new_ws.title = "字表"

    # 创建表头，第1列是'characters'，后续列根据文件名称生成
    headers = ['characters'] + file_names
    new_ws.append(headers)

    # 填充数据
    for row_idx, char in enumerate(reference_chars, start=2):  # 从第2行开始填充数据
        row_data = [char] + merged_data[char]
        new_ws.append(row_data)

        # 为每个文件列添加批注
        for col_idx in range(2, len(file_names) + 2):  # 从第2列开始到最后
            cell = new_ws.cell(row=row_idx, column=col_idx)
            file_index = col_idx - 2  # 计算在文件列表中的索引
            if comments_data[char][file_index]:
                comment_text = "; ".join(comments_data[char][file_index])
                if cell.comment:
                    # 如果已经有批注，保留原来的批注并追加新的内容
                    original_comment = cell.comment.text
                    new_comment_text = f"{original_comment}\n{comment_text}"
                    cell.comment = Comment(new_comment_text, "Python Script")
                else:
                    # 如果没有批注，直接添加新批注
                    cell.comment = Comment(comment_text, "Python Script")

    return new_wb
