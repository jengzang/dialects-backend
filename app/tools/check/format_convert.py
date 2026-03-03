# -*- coding: utf-8 -*-

"""
本腳本將整合音典、跳跳老鼠、縣志三種格式的完整字表提取邏輯，
支援 .tsv、.xlsx、.xls、.docx 格式
根據預設表或用戶選擇對應格式，轉換為 #漢字 音標 解釋 的 .tsv 文件。
"""

import csv
import math
import os
import re
from itertools import product

import docx
import pandas as pd
from docx import Document
from docx.table import Table
from docx.text.paragraph import Paragraph
from openpyxl import load_workbook
from xlrd import open_workbook

from app.common.constants import col_map, vowel_pattern
from app.common.s2t import s2t_pro

# 预编译正则表达式 - 性能优化
RE_SYMBOLS = re.compile(r'[？?＊*]')
RE_CHINESE = re.compile(r'[\u4e00-\u9fa5]')
RE_VOWEL_FALLBACK = re.compile(r"([ʐɣmnŋɲȵƞʋvʒlḷfzr])")
RE_DIGIT = re.compile(r'\d')
RE_TONE_MATCH = re.compile(r"([0-9¹²³⁴⁵⁶⁷⁸⁹⁰]{1,4}[A-Da-d]?)$")
RE_PROCESS_LINES_SUB1 = re.compile(r":\[")
RE_PROCESS_LINES_SUB2 = re.compile(r"\[(\d+)\]")
RE_PROCESS_LINES_SUB3 = re.compile(r"［([^\d]+.*?)］")
RE_COUNTY_MATCHES = re.compile(r"[［\[](\d+[a-z]?)[］\]](.+?)(?=([［\[]\d|$))")
RE_COUNTY_BRACE = re.compile(r"[{｛]([^{}｛｝]+)[}｝]")
RE_VOWEL_RHYME = re.compile(rf".*?([ʐzflḷɣrmnŋɲȵƞʋvʒ].*?)(?=\d|\s|$)")
RE_VOWEL_PATTERN_COMP = re.compile(vowel_pattern)
RE_CHINESE_CHECK = re.compile(r'[一-鿿]')
RE_DIGIT_START = re.compile(r'''^[\d\/?\'"、|；：，。:;,.]+$''')


# def get_tsv_name(path):
#     return os.path.splitext(path)[0] + ".tsv"
# print(docx.__version__)
def get_tsv_name(xls):
    name = os.path.basename(xls)
    name = re.sub(r" ?(\(\d{0,3}\))+$", "", name.rsplit(".", 1)[0]) + ".tsv"
    return os.path.join(os.path.dirname(__file__), name)


def xls_to_tsv(xls_path, page=0):
    def is_xls(fname):
        return fname.endswith("xls") or fname.endswith("xlsx")

    def process_fs(v):
        t = type(v)
        if t is float or t is int:
            return "%d" % v
        if v is None:
            return ""
        return str(v).strip().replace("\t", " ").replace("\n", " ")

    def process_xlsx_fs(v):
        t = type(v)
        if t is float or t is int:
            return "%d" % v
        if v is None:
            return ""
        if t is str:
            return v.strip().replace("\t", " ").replace("\n", " ")
        cells = []
        for i in v:
            if isinstance(i, str):
                cells.append(i.strip())
                continue
            if isinstance(i, (int, float)):
                cells.append("%d" % i)
                continue
            text = i.text
            tag = ""
            if i.font.underline == "single":
                tag = "-"
            elif i.font.underline == "double":
                tag = "="
            if tag:
                text = "".join([j + tag for j in text])
            if i.font.vertAlign == "subscript" or (i.font.size and i.font.size < 10.0):
                text = f"({text})"
            cells.append(text)
        return "".join(cells).replace(")(", "").strip()

    def get_tsv_name(path):
        return os.path.splitext(path)[0] + ".tsv"

    print(f"[INFO] Starting conversion: {xls_path}")
    if not os.path.exists(xls_path):
        print(f"[ERROR] File does not exist: {xls_path}")
        return

    tsv_path = get_tsv_name(xls_path)
    print(f"[INFO] Target TSV path: {tsv_path}")

    lines = []
    header_written = False
    num_columns = 0

    if xls_path.endswith(".xlsx"):
        print("[INFO] Detected .xlsx file")
        wb = load_workbook(xls_path, data_only=True, rich_text=True)
        sheet = wb.worksheets[page]
        print(f"[INFO] Loaded worksheet: {sheet.title}")
        for row_idx, row in enumerate(sheet.rows):
            cols = [process_xlsx_fs(cell.value) for cell in row[:50]]
            if any(cols):
                if not header_written:
                    num_columns = len(cols)
                    header_written = True
                cols += [""] * (num_columns - len(cols))
                lines.append("\t".join(cols[:num_columns]) + "\n")
    else:
        print("[INFO] Detected .xls file")
        wb = open_workbook(xls_path)
        sheet = wb.sheet_by_index(page)
        print(f"[INFO] Loaded sheet: {sheet.name}")
        for i in range(sheet.nrows):
            row = sheet.row_values(i)
            cols = [process_fs(cell) for cell in row]
            if any(cols):
                if not header_written:
                    num_columns = len(cols)
                    header_written = True
                cols += [""] * (num_columns - len(cols))
                lines.append("\t".join(cols[:num_columns]) + "\n")

    print(f"[INFO] Writing {len(lines)} rows to TSV")
    with open(tsv_path, "w", encoding="utf-8", newline="\n") as f:
        f.writelines(lines)

    print(f"[INFO] Conversion complete: {tsv_path}")
    return tsv_path


def run2text(run):
    if isinstance(run, docx.text.hyperlink.Hyperlink):
        return "".join(map(run2text, run.runs))
    tag = ""
    if run.font.underline == docx.enum.text.WD_UNDERLINE.SINGLE:
        tag = "-"
    elif run.font.underline == docx.enum.text.WD_UNDERLINE.DOUBLE:
        tag = "="
    elif run.font.underline == docx.enum.text.WD_UNDERLINE.WAVY:
        tag = chr(0x1AB6)
    elif run._r.xpath("*/w:em[@w:val='dot']"):
        tag = chr(0x0323)
    text = run.text
    if tag:
        text = "".join([i + tag for i in text])
    if run.font.subscript or (run.font.size and run.font.size < docx.shared.Pt(9)):
        if text.startswith("{") and text.endswith("}"):
            pass
        elif text.startswith("[") and text.endswith("]"):
            pass
        else:
            text = f"{{{text}}}"
    return text


def docx_to_tsv(doc):
    if not os.path.exists(doc):
        print("❌ 輸入檔案不存在，跳過")
        return

    lines = []
    Doc = Document(doc)

    for idx, each in enumerate(Doc._body._element):
        if isinstance(each, docx.oxml.table.CT_Tbl):
            # print(f"[📐] 處理表格 第 {idx + 1} 區塊")
            t = Table(each, Doc)
            for row_num, row in enumerate(t.rows):
                行 = ""
                cells = row.cells
                for i, cell in enumerate(cells):
                    if cell in cells[:i]: continue
                    for p in cell.paragraphs:
                        raw = "".join(map(run2text, p.iter_inner_content()))
                        行 += raw.replace("\t", "").replace("\n", "")
                行 += "\t"

                before = 行
                after = 行.replace("}~", "~}").replace("~{", "{~").replace("}{", "").replace("[}", "}[").replace("{h}",
                                                                                                                 "h").strip()
                # print(f"  [→ 表格第 {row_num + 1} 行] 原始：{before}")
                # print(f"  [✓ 清洗後]：{after}")
                lines.append(after)

        elif isinstance(each, docx.oxml.text.paragraph.CT_P):
            # print(f"[📄] 處理段落 第 {idx + 1} 區塊")
            element = Paragraph(each, Doc)
            raw = "".join(map(run2text, element.iter_inner_content()))
            before = raw
            after = raw.replace("}~", "~}").replace("~{", "{~").replace("}{", "").replace("[}", "}[").replace("{h}",
                                                                                                              "h")
            # print(f"  [→ 段落原始]：{before}")
            # print(f"  [✓ 清洗後]：{after}")
            lines.append(after)
    行 = "\n".join(lines).replace("}\n{", "").replace("\n}", "}\n")
    # print(f"[📦] 合併所有行後內容 頭 200 字預覽：\n{行[:200]}...")
    dirpath = os.path.dirname(doc)
    basename = os.path.splitext(os.path.basename(doc))[0]
    tsv_path = os.path.join(dirpath, basename + ".tsv")
    with open(tsv_path, "w", encoding="utf-8", newline="\n") as t:
        t.write(行)
        print(f"[✅] 已寫入：{tsv_path}")

    return tsv_path


def convert_to_tsv_if_needed(filepath):
    ext = os.path.splitext(filepath)[1].lower()
    if ext in [".xlsx", ".xls"]:
        # wb = load_workbook(filepath, data_only=True)
        # sheet = wb.active
        # lines = []
        # for row in sheet.iter_rows(values_only=True):
        #     if all(cell is None for cell in row): continue
        #     line = "\t".join([str(cell).strip() if cell is not None else "" for cell in row])
        #     lines.append(line + "\n")
        # tsv_path = get_tsv_name(filepath)
        # with open(tsv_path, "w", encoding="utf-8") as f:
        #     f.writelines(lines)
        # return tsv_path
        return xls_to_tsv(filepath)
    elif ext == ".docx":
        return docx_to_tsv(filepath)
    else:
        return filepath


# ========== 音典格式處理 ==========
def process_音典(file, level=1, output_path=None):
    print(f"[開始] 處理檔案：{file}")

    # cc = OpenCC('s2t' if level == 1 else 't2s')
    # def s2t(text):
    #     return cc.convert(text)

    file = convert_to_tsv_if_needed(file)
    print(f"[轉換] 轉為 TSV 路徑：{file}")

    rows = []
    simplified_rows = []

    with open(file, encoding="utf-8") as f:
        lines = [line.rstrip("\n").split("\t") for line in f if line.strip() and not line.startswith("#")]

    if not lines:
        # print("⚠️ 無有效數據，檔案內容為空或格式錯誤")
        # with open(WRITE_ERROR_LOG, "a", encoding="utf-8") as f:
        #     f.write(f"⚠️ [{file}] 無有效數據，檔案內容為空或格式錯誤\t【format_convert->process_音典】\n")
        return

    header = lines[0]
    print(f"[分析] 表頭：{header}")

    index = {}
    for std_key, aliases in col_map.items():
        for i, name in enumerate(header):
            if name.strip().lower() in [a.lower() for a in aliases]:
                index[std_key] = i
                print(f"✅ 欄位對應：{std_key} → 第 {i + 1} 欄（{name}）")
                break

    if '漢字' not in index or '音標' not in index:
        print("❌ 欄位對應失敗，請確認有『漢字』與『音標』欄位")
        return

    print(f"[處理] 開始掃描資料行，共 {len(lines) - 1} 筆")

    delimiters = [';', '；', '/', '、', ',', '，']

    def split_field(field):
        for delim in delimiters:
            field = field.replace(delim, '∥')  # 統一分隔符為 ∥
        return [f.strip() for f in field.split('∥') if f.strip()]

    def get_field(parts, field_name):
        idx = index.get(field_name)
        if idx is not None and idx < len(parts):
            return parts[idx].strip()
        return ""

    def process_pair(word, phon, note, row_num, level):
        if level == 0:
            clean_str, mapping = word, {}  # 伪代码：跳过转换
        else:
            clean_str, mapping = s2t_pro(word, level)
        mapping = dict(mapping)
        phon_units = phon.strip().split()
        word_len_match = len(word) == len(phon_units)

        if word_len_match:
            for ch, p in zip(word, phon_units):
                candidates = mapping.get(ch, [ch])
                for cand in candidates:
                    rows.append([cand, p, note])
                    if cand != ch:
                        simplified_rows.append([cand, p, note, "簡"])
                        print(f"[簡體一對多] 第 {row_num} 行：{ch} → {cand}")
        else:
            rows.append([clean_str, phon, note])
            if clean_str != word:
                simplified_rows.append([clean_str, phon, note, "簡"])
                print(f"[fallback] 第 {row_num} 行：{word} → {clean_str}")

    # print(f"[處理] 開始掃描資料行，共 {len(lines) - 1} 筆")

    for row_num, parts in enumerate(lines[1:], start=2):
        word_raw = get_field(parts, '漢字')
        phon_raw = get_field(parts, '音標')
        note = get_field(parts, '解釋')

        if not word_raw or not phon_raw:
            continue

        word_list = split_field(word_raw)
        phon_list = split_field(phon_raw)

        if not word_list or not phon_list:
            print(f"⚠️ 跳過第 {row_num} 行，因為漢字或音標清單為空")
            continue

        if len(word_list) > 1 and len(phon_list) > 1:
            # ✅ 無論等長與否，始終做笛卡爾積
            print(f"[笛卡爾積] 第 {row_num} 行：{word_list} × {phon_list}")
            for word, phon in product(word_list, phon_list):
                process_pair(word, phon, note, row_num, level)

        elif len(word_list) > 1 and len(phon_list) == 1:
            # ✅ 多對一
            print(f"[多對一] 第 {row_num} 行：{word_list} × {phon_list[0]}")
            for word in word_list:
                process_pair(word, phon_list[0], note, row_num, level)

        elif len(word_list) == 1 and len(phon_list) > 1:
            # ✅ 一對多
            print(f"[一對多] 第 {row_num} 行：{word_list[0]} × {phon_list}")
            for phon in phon_list:
                process_pair(word_list[0], phon, note, row_num, level)

        else:
            # fallback 合併處理
            word = ''.join(word_list)
            phon = ' '.join(phon_list)
            # print(f"[fallback] 第 {row_num} 行：{word} → {phon}")
            process_pair(word, phon, note, row_num, level)

    outpath = output_path or (os.path.splitext(file)[0] + ".tsv")
    print(f"[輸出] 寫入主檔案：{outpath}")

    with open(outpath, "w", encoding="utf-8", newline="\n") as out:
        writer = csv.writer(out, delimiter="\t")
        writer.writerow(["#漢字", "音標", "解釋"])
        writer.writerows(rows)

    # simp_path = os.path.splitext(file)[0] + ".簡.tsv"
    # if simplified_rows:
    #     print(f"[簡體] 共發現 {len(simplified_rows)} 筆簡體詞彙，寫入：{simp_path}")
    #     with open(simp_path, "w", encoding="utf-8", newline="\n") as out:
    #         writer = csv.writer(out, delimiter="\t")
    #         writer.writerow(["#漢字", "音標", "解釋", "繁簡"])
    #         writer.writerows(simplified_rows)

    print(f"✅ 全部處理完成：{outpath}")


# ========== 跳跳老鼠格式處理 ==========
def process_跳跳老鼠(file, level=1, output_path=None):
    print(f"📄 開始處理文件：{file}")
    rows = []
    simplified_rows = []

    # 選擇繁→簡或簡→繁
    # converter = OpenCC('s2t' if level == 1 else 't2s')
    # def s2t(text):
    #     return converter.convert(text)

    # 讀取 Excel（僅第一張表）
    wb = load_workbook(file, data_only=True)
    sheet = wb.active

    def parse_row(line, line_num):
        parts = [str(c).strip() if c is not None else "" for c in line]
        if len(parts) < 2:
            print(f"⚠️ 第 {line_num} 行欄位不足，跳過：{parts}")
            return []
        phon = parts[0]
        組 = parts[1]
        if not phon or not 組:
            print(f"⚠️ 第 {line_num} 行缺音或字，跳過")
            return []
        result = []
        matches = re.findall(r"(.)(?:\{(.*?)\}|\[(.*?)\])?", 組)
        # print(f"🔍 第 {line_num} 行組拆分：{matches}")
        for 字, 註1, 註2 in matches:
            註 = 註1 or 註2 or ""
            # print(f"🧩 字：{字}，音：{phon}，註：{註}")
            result.append((字, phon, 註))
        return result

    for i, row in enumerate(sheet.iter_rows(values_only=True), start=1):
        if not row or str(row[0]).startswith("#"):
            continue
        parsed = parse_row(row, i)
        for 字, 音, 註 in parsed:
            if level == 0:
                clean_str, mapping = 字, {}  # 伪代码：跳过转换
            else:
                clean_str, mapping = s2t_pro(字, level)

            mapping = dict(mapping)
            candidates = mapping.get(字, [字])  # 支援多候選

            for cand in candidates:
                rows.append([cand, 音, 註])
                if cand != 字:
                    simplified_rows.append([cand, 音, 註, "簡"])
                    # print(f"🔁 字形轉換：{字} → {cand}")

    outpath = output_path or os.path.splitext(file)[0] + ".tsv"
    with open(outpath, "w", encoding="utf-8", newline="\n") as out:
        writer = csv.writer(out, delimiter="\t")
        writer.writerow(["#漢字", "音標", "解釋"])
        writer.writerows(rows)
    print(f"✅ 主檔輸出完成：{outpath}")

    simp_path = os.path.splitext(file)[0] + ".簡.tsv"
    if simplified_rows:
        with open(simp_path, "w", encoding="utf-8", newline="\n") as out:
            writer = csv.writer(out, delimiter="\t")
            writer.writerow(["#漢字", "音標", "解釋", "繁簡"])
            writer.writerows(simplified_rows)
        print(f"[簡體] 共發現 {len(simplified_rows)} 筆簡體詞彙，寫入：{simp_path}")

    print(f"🎉 全部處理完成，共 {len(rows)} 條記錄")
    # print(f"✅ 輸出：{outpath}")


# ========== 縣志格式處理 ==========
def process_縣志_excel(file, level=1, output_path=None):
    # cc = OpenCC('s2t')
    rows = []
    simplified_rows = []
    debug = False

    # def s2t(text, level=1):
    #     return cc.convert(text)

    def process_lines(行):
        行 = 行.strip()
        if not 行:
            return None
        if 行.startswith("#"):
            return 行
        行 = RE_PROCESS_LINES_SUB1.sub("\t[", 行)
        行 = 行.replace("(", "{").replace(")", "}")
        行 = RE_PROCESS_LINES_SUB2.sub(r"［\1］", 行)
        行 = RE_PROCESS_LINES_SUB3.sub(r"[\1]", 行)
        return 行

    ext = os.path.splitext(file)[1].lower()
    if ext in [".xlsx", ".xls"]:
        df = pd.read_excel(file, sheet_name=0, header=None)
        lines = [
            "\t".join([str(cell) for cell in row if pd.notna(cell)]).strip()
            for _, row in df.iterrows()
        ]
        print(f"📖 讀取 Excel：{file}")
    else:
        encodings = ["utf-8", "utf-8-sig", "big5", "gb18030"]
        for enc in encodings:
            try:
                with open(file, encoding=enc) as f:
                    lines = f.readlines()
                print(f"📖 使用編碼：{enc}")
                break
            except UnicodeDecodeError:
                continue
        else:
            raise UnicodeDecodeError("❌ 無法讀取文件，請確認編碼格式")

    total, skipped, simplified_count = 0, 0, 0

    for lineno, line in enumerate(lines, 1):
        total += 1
        raw_line = line
        line = process_lines(line)
        if line is None:
            skipped += 1
            continue

        if line.startswith("#漢字"):
            skipped += 1
            continue
        if line.startswith("#"):
            continue

        parts = line.split("\t")
        if len(parts) < 2:
            if debug:
                print(f"⚠️ 跳過行 {lineno}（分欄不足）: {raw_line.strip()}")
            skipped += 1
            continue

        拼音 = parts[0].strip()
        for cell in parts[1:]:
            matches = RE_COUNTY_MATCHES.findall(cell)
            if not matches:
                if debug:
                    print(f"⚠️ 無音節匹配 行 {lineno}: {cell}")
                continue

            for 調號, 義項, _ in matches:
                if debug:
                    print(f"🔎 行 {lineno}：拼音={拼音}, 調號={調號}, 義項={義項}")

                # 逐字掃描義項，若某字後緊跟註釋，就綁定在那個字上
                i = 0
                while i < len(義項):
                    字 = 義項[i]
                    註 = ""
                    if i + 1 < len(義項) and 義項[i + 1] in "{｛":
                        m = RE_COUNTY_BRACE.match(義項[i + 1:])
                        if m:
                            註 = m.group(1)
                            i += len(m.group(0))  # 跳過整個 {註釋}
                    i += 1
                    字 = 字.strip()
                    if not 字:
                        if debug:
                            print(f"⚠️ 空白字 行 {lineno} 義項：{義項}")
                        continue

                    if level == 0:
                        clean_str, mapping = 字, {}  # 伪代码：跳过转换
                    else:
                        clean_str, mapping = s2t_pro(字, level)

                    mapping = dict(mapping)
                    candidates = mapping.get(字, [字])  # 支援多候選繁體字

                    音標 = f"{拼音}{調號}"
                    for cand in candidates:
                        row = [cand, 音標, 註]
                        rows.append(row)
                        if cand != 字:
                            simplified_rows.append(row + ["簡"])
                            simplified_count += 1
                            if debug:
                                print(f"🔁 字形轉換：{字} → {cand}")

    outpath = output_path or os.path.splitext(file)[0] + ".tsv"
    with open(outpath, "w", encoding="utf-8", newline="\n") as out:
        writer = csv.writer(out, delimiter="\t")
        writer.writerow(["#漢字", "音標", "解釋"])
        writer.writerows(rows)
    print(f"✅ 主檔輸出完成：{outpath}")

    simp_path = os.path.splitext(file)[0] + ".簡.tsv"
    if simplified_rows:
        with open(simp_path, "w", encoding="utf-8", newline="\n") as out:
            writer = csv.writer(out, delimiter="\t")
            writer.writerow(["#漢字", "音標", "解釋", "繁簡"])
            writer.writerows(simplified_rows)
        print(f"[簡體] 共發現 {len(simplified_rows)} 筆簡體詞彙，寫入：{simp_path}")

    print(f"📊 行數統計：總行數 {total}, 跳過 {skipped} 行, 標註簡體 {simplified_count} 條")


def process_縣志_word(file, level=1, output_path=None):
    print(f"📖 讀取 word：{file}")
    tsv_path = convert_to_tsv_if_needed(file)
    with open(tsv_path, encoding="utf-8") as f:
        raw = f.read()

    def parse_entry_blocks(text):
        results = []

        current_vowel = None  # e.g., 'i', 'u'

        for line in text.splitlines():
            line = line.strip()
            if not line:
                continue

            if line.startswith(("#", "＃")):
                # 去掉第一个字符并取剩余部分作为当前元音
                current_vowel = line[1:].strip()
                continue

            # 1. 匹配开头直到遇到半角 [ 或全角 ［
            match = re.match(r"^([^\[［]+)", line)
            if not match or not current_vowel:
                continue

            initial = match.group(1).strip()

            # 2. 匹配 [数字] 或 ［数字］ 之后的内容
            # 使用 [\[［] 匹配两种左括号，[\]］] 匹配两种右括号
            segments = re.findall(r"[\[［](\d+)[\]］]([^\[［]+)", line)

            for tone, content in segments:
                syllable = f"{initial}{current_vowel}{tone}"
                chars = []
                explanations = {}
                temp = ""
                in_brace = False
                current_char = ""

                for c in content:
                    if c in ("{", "｛"):
                        in_brace = True
                        temp = ""
                    elif c in ("}", "｝"):
                        in_brace = False
                        explanations[current_char] = temp
                        temp = ""
                    elif in_brace:
                        temp += c
                    else:
                        current_char = c
                        chars.append(c)

                for char in chars:
                    explanation = explanations.get(char, "")
                    results.append((char, syllable, explanation))

        return results

    data = parse_entry_blocks(raw)
    outpath = output_path or os.path.splitext(file)[0] + ".tsv"
    df = pd.DataFrame(data, columns=["#漢字", "音標", "解釋"])
    df.to_csv(outpath, sep="\t", index=False)
    # print(f"[✅] 轉換完成，輸出路徑：{outpath}")

    with open(outpath, encoding="utf-8") as f:
        lines = [line.rstrip("\n").split("\t") for line in f if line.strip() and not line.startswith("#")]
        # print("lines:",lines)

    index = {'漢字': 0, '音標': 1, '解釋': 2}
    print(f"[處理] 開始掃描資料行，共 {len(lines) - 1} 筆")

    delimiters = [';', '；', '/', '、', ',', '，']

    rows = []
    simplified_rows = []

    def split_field(field):
        for delim in delimiters:
            field = field.replace(delim, '∥')
        return [f.strip() for f in field.split('∥') if f.strip()]

    def get_field(parts, field_name):
        idx = index.get(field_name)
        if idx is not None and idx < len(parts):
            return parts[idx].strip()
        return ""

    def process_pair(word, phon, note, row_num, level):
        if level == 0:
            clean_str, mapping = word, {}  # 伪代码：跳过转换
        else:
            clean_str, mapping = s2t_pro(word, level)

        mapping = dict(mapping)
        phon_units = phon.strip().split()
        word_len_match = len(word) == len(phon_units)

        if word_len_match:
            for ch, p in zip(word, phon_units):
                candidates = mapping.get(ch, [ch])
                for cand in candidates:
                    rows.append([cand, p, note])
                    if cand != ch:
                        simplified_rows.append([cand, p, note, "簡"])
                        # print(f"[簡體一對多] 第 {row_num} 行：{ch} → {cand}")
        else:
            rows.append([clean_str, phon, note])
            if clean_str != word:
                simplified_rows.append([clean_str, phon, note, "簡"])
                # print(f"[fallback] 第 {row_num} 行：{word} → {clean_str}")

    print(f"[處理] 開始掃描資料行，共 {len(lines) - 1} 筆")

    for row_num, parts in enumerate(lines[1:], start=2):
        word_raw = get_field(parts, '漢字')
        phon_raw = get_field(parts, '音標')
        note = get_field(parts, '解釋')

        if not word_raw or not phon_raw:
            continue

        word_list = split_field(word_raw)
        phon_list = split_field(phon_raw)

        if not word_list or not phon_list:
            print(f"⚠️ 跳過第 {row_num} 行，因為漢字或音標清單為空")
            continue

        if len(word_list) > 1 and len(phon_list) > 1:
            # print(f"[笛卡爾積] 第 {row_num} 行：{word_list} × {phon_list}")
            for word, phon in product(word_list, phon_list):
                process_pair(word, phon, note, row_num, level)

        elif len(word_list) > 1 and len(phon_list) == 1:
            # print(f"[多對一] 第 {row_num} 行：{word_list} × {phon_list[0]}")
            for word in word_list:
                process_pair(word, phon_list[0], note, row_num, level)

        elif len(word_list) == 1 and len(phon_list) > 1:
            # print(f"[一對多] 第 {row_num} 行：{word_list[0]} × {phon_list}")
            for phon in phon_list:
                process_pair(word_list[0], phon, note, row_num, level)

        else:
            word = ''.join(word_list)
            phon = ' '.join(phon_list)
            # print(f"[fallback] 第 {row_num} 行：{word} → {phon}")
            process_pair(word, phon, note, row_num, level)

    # Step 5: 輸出最終 TSV
    # final_outpath = output_path or (os.path.splitext(file)[0] + ".tsv")
    with open(outpath, "w", encoding="utf-8", newline="\n") as out:
        writer = csv.writer(out, delimiter="\t")
        writer.writerow(["#漢字", "音標", "解釋"])
        writer.writerows(rows)

    # Step 6: 輸出簡體資料
    simp_path = os.path.splitext(file)[0] + ".簡.tsv"
    if simplified_rows:
        print(f"[簡體] 共發現 {len(simplified_rows)} 筆簡體詞彙，寫入：{simp_path}")
        with open(simp_path, "w", encoding="utf-8", newline="\n") as out:
            writer = csv.writer(out, delimiter="\t")
            writer.writerow(["#漢字", "音標", "解釋", "繁簡"])
            writer.writerows(simplified_rows)

    print(f"✅ 全部處理完成：{outpath}")


def process_縣志(file, level=1, output_path=None):
    ext = os.path.splitext(file)[1].lower()
    if ext in [".xlsx", ".xls"]:
        process_縣志_excel(file, level, output_path)
    elif ext in [".docx", ".doc"]:
        process_縣志_word(file, level, output_path)


# 1. 将核心提取逻辑提取为独立工具函数，供两者调用
def _core_extract_logic(phon: str, tone_map: dict = None) -> tuple[str, str, str]:
    """核心提取逻辑：处理单个音标字符串，返回 (声母, 韵母, 声调)"""

    phon = phon.strip()
    if not phon: return "", "", ""

    # --- 声母提取 ---
    consonant = ""
    if phon[0] in {"∅", "Ø", "0"}:
        consonant = "ʔ"
    else:
        pre_digit_part = RE_DIGIT.split(phon)[0]
        if not RE_VOWEL_PATTERN_COMP.search(pre_digit_part):
            if RE_VOWEL_FALLBACK.match(phon[0]):
                consonant = "/"
            elif not RE_VOWEL_FALLBACK.search(phon):
                consonant = ""
            else:
                for char in phon:
                    if RE_VOWEL_FALLBACK.match(char) or RE_DIGIT.match(char): break
                    consonant += char
        else:
            if RE_VOWEL_PATTERN_COMP.match(phon[0]):
                consonant = "/"
            elif 'j' in phon[1:] or 'ʲ' in phon[1:]:
                for char in phon:
                    if RE_VOWEL_PATTERN_COMP.match(char) or char in ('j', 'ʲ'): break
                    consonant += char
            else:
                for char in phon:
                    if RE_VOWEL_PATTERN_COMP.match(char): break
                    consonant += char
        consonant = RE_DIGIT.sub("", consonant)

    # --- 韵母提取 ---
    all_rhymes = []
    tmp_phon = phon[1:] if phon.startswith(("∅", "Ø", "0")) else phon
    if 'j' not in tmp_phon[1:] and 'ʲ' not in tmp_phon[1:]:
        vowel_found = False
        for c in tmp_phon:
            if RE_VOWEL_PATTERN_COMP.match(c) and not vowel_found:
                vowel_found = True
                all_rhymes.append(c)
            elif vowel_found and (c.isdigit() or c.isspace()):
                break
            elif vowel_found:
                all_rhymes.append(c)
        if not vowel_found and any(c in tmp_phon for c in "ʐzflḷɣmnŋȵɲƞʋvʒr"):
            match = RE_VOWEL_RHYME.search(tmp_phon)
            if match: all_rhymes += list(match.group(1))
    else:
        match = re.search(rf"[{vowel_pattern.strip('[]')}jʲ][^\d\s]*", tmp_phon)
        if match: all_rhymes = list(match.group(0))

    rhyme = ''.join(c for c in all_rhymes if not (c.isdigit() or RE_CHINESE_CHECK.match(c)))

    # --- 标准化替换 ---
    for old, new in {'ε': 'ɛ', "α": "ɑ", "ʯ": "ʮ", "∅": "ø", "ο": "o", "ǝ": "ə", "о": "o", "у": "y", "е": "e", "ã": "ã",
                     "ẽ": "ẽ", "ĩ": "ĩ", "ī": "ĩ", "ā": "ã", "ỹ": "ỹ", "õ": "õ", "ʱ": "ʰ"}.items():
        rhyme = rhyme.replace(old, new)
    for old, new in {'∫': 'ʃ', 'th': 'tʰ', 'kh': 'kʰ', 'ph': 'pʰ', 'tsh': 'tsʰ', "ς": "ɕ", 'ts': 'ʦ', 'tʃ': 'ʧ',
                     'tɕ': 'ʨ', "∨": "v", "ł": "ɬ", "tʰs": "ʦʰ", "(ʔ)": "ʔ", "∅": "ʔ", "Ǿ": "ʔ"}.items():
        consonant = consonant.replace(old, new)

    # --- 声调提取 ---
    tone = ""
    if "輕聲" in phon:
        tone = "輕聲"
    else:
        # 推薦的最小化改動：直接匹配末尾
        tone_match = RE_TONE_MATCH.search(phon.strip())
        if tone_match:
            tone_code = tone_match.group(1)
            # 如果有 tone_map 就查表，沒有就直接用提取到的 code
            tone = tone_map.get(tone_code, tone_code) if tone_map else tone_code
        else:
            tone = ""
    return consonant, rhyme, tone


def extract_all_from_files(file_path: str, preserve_empty_rows: bool = True) -> pd.DataFrame:
    def get_standard_column_name(col_name, col_map):
        """
        根據 col_map 返回標準化的列名。
        :param col_name: 當前列名
        :param col_map: 列名映射
        :return: 返回對應的標準化列名
        """
        for standard_col, possible_names in col_map.items():
            if col_name in possible_names:
                return standard_col
        return col_name  # 如果找不到對應的列名，返回原列名

    # 檢查文件的副檔名來決定使用哪種方法
    file_extension = os.path.splitext(file_path)[1].lower()
    # print(file_extension)
    if file_extension == ".tsv":
        df = pd.read_csv(file_path, sep="\t", dtype=str)
    elif file_extension in [".xls", ".xlsx"]:
        df = pd.read_excel(file_path, dtype=str)
    else:
        raise ValueError("Unsupported file format. Please provide a TSV or Excel file.")

        # 處理欄位名稱，根據 col_map 進行模糊對應
    df.columns = [get_standard_column_name(col, col_map) for col in df.columns]
    df = df.fillna("")

    results = []
    df = df.loc[:, ~df.columns.duplicated()]
    for _, row in df.iterrows():
        hanzi = str(row.get("漢字", "")).strip()
        phonetic = str(row.get("音標", "")).strip()
        note = str(row.get("解釋", "")).strip()
        if isinstance(phonetic, str):
            phonetic = phonetic.strip()
        if not hanzi or not phonetic or phonetic == "0" or hanzi == "0":
            if preserve_empty_rows:
                results.append({
                    '汉字': '',
                    '音标': '',
                    '声母': '',
                    '韵母': '',
                    '声调': '',
                    '註釋': ''
                })
                continue
            continue
        # if re.search(r"[□■⬜⬛☐☑☒▯▢▣█�]", hanzi):
        #     if preserve_empty_rows:
        #         results.append({
        #             '汉字': '',
        #             '音标': '',
        #             '声母': '',
        #             '韵母': '',
        #             '声调': '',
        #             '註釋': ''
        #         })
        #         continue
        #     continue
        if phonetic and (isinstance(phonetic, float) and math.isnan(phonetic)):
            if preserve_empty_rows:
                results.append({
                    '汉字': '',
                    '音标': '',
                    '声母': '',
                    '韵母': '',
                    '声调': '',
                    '註釋': ''
                })
                continue
            continue
        if isinstance(phonetic, str) and phonetic and phonetic[0].isdigit() and phonetic[0] != '0':
            if preserve_empty_rows:
                results.append({
                    '汉字': '',
                    '音标': '',
                    '声母': '',
                    '韵母': '',
                    '声调': '',
                    '註釋': ''
                })
                continue
            continue

        phonetic_variants = phonetic.split("/") if "/" in phonetic and phonetic.strip() != '/' else [phonetic]

        for phon in phonetic_variants:
            phon = phon.strip()  # 先清乾淨
            if not phon:  # 若為空，跳過（這一步是關鍵防炸）
                continue
            if RE_DIGIT_START.match(phon):
                continue

            # 调用核心逻辑
            consonant, rhyme, tone = _core_extract_logic(phon)

            results.append({
                '汉字': hanzi,
                '音标': phon,
                '声母': consonant,
                '韵母': rhyme,
                '声调': tone,
                '註釋': note
            })
    return pd.DataFrame(results)

# 3. 简化的 extract_onset_rime_from_ipa
def extract_onset_rime_from_ipa(ipa: str) -> tuple[str, str, str]:
    """直接调用核心逻辑，不传 tone_map 则返回原始调值"""
    return _core_extract_logic(ipa)
